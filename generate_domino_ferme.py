#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère les fichiers Excel pour les abris Domino FERMÉS
- Murs : partout (haut, droite, bas, gauche)
- Portes et serrure : pré-configurées dans le fichier de base
- Remove cladding : Yes (B25 = Yes)
- Variantes : normaux et bosqués
- Chaque variante : Galvanized/Powder coated × Standard/PLUS
- Toutes les combinaisons de largeurs et profondeurs
"""

import openpyxl
import shutil
import os
import json
from datetime import datetime

# Dossier et fichier source
base_dir = 'fichier de base'
source_file = os.path.join(base_dir, 'nepastoucher.xlsx')
resultats_dir = 'résultats'

# Valeurs valides depuis Excel (feuille Calc)
traitements = ['Galvanized', 'Powder coated']
versions = ['Standard', 'PLUS']

# Variantes - Pour Domino, remove_cladding = Yes
variantes = [
    {'nom': 'normal', 'wall_material': 'Thermowood', 'remove_cladding': 'Yes'},
    {'nom': 'bosque', 'wall_material': 'Thermowood', 'remove_cladding': 'Yes'}
]

# Fonction pour décomposer la profondeur en valeurs valides (2m, 2.5m)
def decomposer_profondeur(profondeur_totale):
    """
    Décompose une profondeur totale en valeurs valides (2.03m, 2.53m)
    Retourne une liste de valeurs à mettre dans A2, A3, A4, etc.
    Règles exactes pour les FERMÉS :
    - 4m = 2m + 2m
    - 4.5m = 2m + 2.5m
    - 5m = 2.5m + 2.5m
    - 6m = 2m + 2m + 2m
    - 7m = 2m + 2.5m + 2.5m
    - 8m = 2m + 2m + 2m + 2m
    - 9m = 2.5m + 2m + 2m + 2.5m (symétrique)
    - 10m = 2m + 2m + 2m + 2m + 2m (5 fois 2)
    - 11m = 2.5m + 2m + 2m + 2m + 2m (2.5 au début puis que des 2)
    - 12m = 2m + 2m + 2m + 2m + 2m + 2m (6 fois 2)
    """
    # Cas spéciaux selon les règles
    if profondeur_totale == 4:
        return [2.03, 2.03]
    elif profondeur_totale == 4.5:
        return [2.03, 2.53]
    elif profondeur_totale == 5:
        return [2.53, 2.53]
    elif profondeur_totale == 6:
        return [2.03, 2.03, 2.03]
    elif profondeur_totale == 7:
        return [2.03, 2.53, 2.53]
    elif profondeur_totale == 8:
        return [2.03, 2.03, 2.03, 2.03]  # 4 fois 2
    elif profondeur_totale == 9:
        return [2.53, 2.03, 2.03, 2.53]  # Symétrique : 2.5 + 2 + 2 + 2.5
    elif profondeur_totale == 10:
        return [2.03, 2.03, 2.03, 2.03, 2.03]  # 5 fois 2
    elif profondeur_totale == 11:
        return [2.53, 2.03, 2.03, 2.03, 2.03]  # 2.5 au début puis que des 2
    elif profondeur_totale == 12:
        return [2.03, 2.03, 2.03, 2.03, 2.03, 2.03]  # 6 fois 2
    else:
        # Pour les autres valeurs, algorithme glouton
        valeurs_valides = [2.03, 2.53]
        resultat = []
        reste = profondeur_totale
        
        while reste > 0.1:
            if reste >= 2.53:
                resultat.append(2.53)
                reste -= 2.53
            elif reste >= 2.03:
                resultat.append(2.03)
                reste -= 2.03
            else:
                resultat.append(2.03)
                reste = 0
        
        return resultat

# Fonction pour décomposer la largeur en valeurs valides (2m, 2.5m, 4m, 5m, 6m)
def decomposer_largeur(largeur_totale):
    """
    Décompose une largeur totale en valeurs valides (2.03, 2.53, 4.06, 5.06, 6.09)
    Retourne une liste de valeurs à mettre dans B1, C1, D1, etc.
    """
    valeurs_valides = [2.03, 2.53, 4.06, 5.06, 6.09]
    
    # Cas spéciaux selon les règles données
    if largeur_totale == 2:
        return [2.03]
    elif largeur_totale == 2.5:
        return [2.53]
    elif largeur_totale == 3:
        return [2.53, 2.03]
    elif largeur_totale == 4:
        return [4.06]
    elif largeur_totale == 4.5:
        return [4.06]
    elif largeur_totale == 5:
        return [5.06]
    elif largeur_totale == 6:
        return [6.09]
    elif largeur_totale == 7:
        return [2.53, 2.03, 2.53]
    elif largeur_totale == 8:
        return [4.06, 4.06]
    elif largeur_totale == 9:
        return [2.53, 4.06, 2.53]
    elif largeur_totale == 10:
        return [5.06, 5.06]
    elif largeur_totale == 11:
        return [2.53, 6.09, 2.53]
    elif largeur_totale == 12:
        return [6.09, 6.09]
    elif largeur_totale == 13:
        return [4.06, 5.06, 4.06]
    elif largeur_totale == 14:
        return [5.06, 4.06, 5.06]
    else:
        # Algorithme glouton pour les autres cas (15m+)
        resultat = []
        reste = largeur_totale
        valeurs_triees = sorted(valeurs_valides, reverse=True)
        
        while reste > 0.1:
            trouve = False
            for val in valeurs_triees:
                if reste >= val:
                    resultat.append(val)
                    reste -= val
                    trouve = True
                    break
            if not trouve:
                resultat.append(valeurs_valides[0])
                reste = 0
        
        return resultat

# Largeurs et profondeurs pour les abris FERMÉS
largeurs_totales = [4, 5, 6, 7, 8]  # 5 largeurs (2 et 2.5 enlevés pour les fermés)
profondeurs_totales = [4, 4.5, 5, 6, 7, 8, 9, 10, 11, 12]  # 10 profondeurs

print("=" * 80)
print("GÉNÉRATION DES ABRIS DOMINO FERMÉS")
print("=" * 80)

# Vérifier que le fichier source existe
if not os.path.exists(source_file):
    print(f"❌ Erreur: {source_file} n'existe pas !")
    exit(1)

# Créer le dossier résultats
os.makedirs(resultats_dir, exist_ok=True)

# Créer un sous-dossier pour les abris Domino fermés
output_dir = os.path.join(resultats_dir, 'domino_ferme')
os.makedirs(output_dir, exist_ok=True)

# Supprimer les anciens fichiers
for old_file in os.listdir(output_dir):
    if old_file.endswith('.xlsx'):
        os.remove(os.path.join(output_dir, old_file))

fichiers_crees = []
compteur = 1

# Générer toutes les combinaisons
for largeur_totale in largeurs_totales:
    for profondeur_totale in profondeurs_totales:
        # Décomposer en valeurs valides
        largeurs_decomposees = decomposer_largeur(largeur_totale)
        profondeurs_decomposees = decomposer_profondeur(profondeur_totale)
        
        for variante in variantes:
            for treatment in traitements:
                for version in versions:
                    # Nom du fichier selon la nomenclature
                    # Format: DOM-F-{largeur}M-{version}-{profondeur}-{treatment}
                    
                    # Type: DOM-F pour Domino fermé
                    type_code = 'DOM-F'
                    
                    # Largeur: format 6M, 10M, etc.
                    largeur_code = f'{int(largeur_totale)}M' if largeur_totale == int(largeur_totale) else f'{largeur_totale}M'
                    
                    # Version: N pour Standard, P pour PLUS
                    version_code = 'N' if version == 'Standard' else 'P'
                    
                    # Profondeur: format 418 pour 4.18m, 621 pour 6.21m, etc.
                    profondeur_cm = int(profondeur_totale * 100)
                    profondeur_code = str(profondeur_cm)
                    
                    # Treatment: G pour Galvanized, PT pour Powder coated
                    treatment_code = 'G' if treatment == 'Galvanized' else 'PT'
                    
                    # Nom du fichier
                    nom_fichier = f'{type_code}-{largeur_code}-{version_code}-{profondeur_code}-{treatment_code}.xlsx'
                    work_file = os.path.join(output_dir, nom_fichier)
                    
                    print(f"\n📦 Création {compteur}: {os.path.basename(work_file)}")
                    print(f"   Largeur totale: {largeur_totale}m → {largeurs_decomposees}")
                    print(f"   Profondeur totale: {profondeur_totale}m → {profondeurs_decomposees}")
                    
                    # Dupliquer le fichier source
                    shutil.copy2(source_file, work_file)
                    
                    # Ouvrir et modifier
                    wb = openpyxl.load_workbook(work_file, data_only=False)
                    ws = wb['Configure']
                    
                    # Nettoyer les lignes 29-31 (supprimer les espaces, mettre à None)
                    for row in range(29, 32):
                        for col in range(1, 4):  # Colonnes A, B, C
                            cell_value = ws.cell(row, col).value
                            if cell_value == ' ' or (isinstance(cell_value, str) and cell_value.strip() == ''):
                                ws.cell(row, col).value = None
                    
                    # Mettre "*" dans toutes les cellules de dimensions
                    for row in range(2, 14):
                        ws.cell(row, 1).value = "*"
                    for col in range(2, 8):
                        ws.cell(1, col).value = "*"
                    
                    # Écrire les profondeurs décomposées (A2, A3, A4, etc.)
                    for i, prof in enumerate(profondeurs_decomposees[:12]):  # Max 12 profondeurs
                        ws.cell(2 + i, 1).value = prof
                    
                    # Écrire les largeurs décomposées (B1, C1, D1, etc.)
                    for i, larg in enumerate(largeurs_decomposees[:6]):  # Max 6 largeurs
                        ws.cell(1, 2 + i).value = larg
                    
                    # Écrire les options de base
                    ws.cell(16, 2).value = treatment  # B16 = treatment
                    ws.cell(17, 2).value = version  # B17 = version
                    
                    # Configuration ABRIS DOMINO FERMÉS
                    ws.cell(19, 2).value = variante['wall_material']  # B19 = wall material
                    ws.cell(21, 2).value = 'Yes'  # B21 = top wall
                    ws.cell(22, 2).value = 'Yes'  # B22 = right wall
                    ws.cell(23, 2).value = 'Yes'  # B23 = bottom wall (FERMÉ)
                    ws.cell(24, 2).value = 'Yes'  # B24 = left wall
                    ws.cell(25, 2).value = 'Yes'  # B25 = remove cladding (YES pour Domino)
                    
                    # Portes pour les fermés - Ajuster B28 selon la profondeur
                    # Si la profondeur ne contient aucune valeur de 2.5m (2.53), mettre 2m (2.03) dans B28
                    # Sinon, garder 2.5m (2.53) dans B28
                    if 2.53 in profondeurs_decomposees:
                        # Il y a au moins un 2.53, donc B28 = 2.53
                        ws.cell(28, 2).value = 2.53
                    else:
                        # Pas de 2.53, donc B28 = 2.03
                        ws.cell(28, 2).value = 2.03
                    
                    # Gate hardware kit pour les fermés
                    # NE RIEN TOUCHER - Le fichier de base est déjà pré-configuré pour les fermés
                    
                    # NE PAS TOUCHER B26 et B27 - ils ne doivent pas être modifiés
                    
                    # Sauvegarder
                    wb.save(work_file)
                    
                    fichiers_crees.append({
                        'fichier': os.path.basename(work_file),
                        'largeur_totale': largeur_totale,
                        'largeurs_decomposees': largeurs_decomposees,
                        'profondeur_totale': profondeur_totale,
                        'profondeurs_decomposees': profondeurs_decomposees,
                        'variante': variante['nom'],
                        'treatment': treatment,
                        'version': version,
                        'type': 'domino_ferme'
                    })
                    
                    compteur += 1

# Sauvegarder le résumé
resume = {
    'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    'type': 'domino_ferme',
    'total_fichiers': len(fichiers_crees),
    'largeurs_totales': largeurs_totales,
    'profondeurs_totales': profondeurs_totales,
    'variantes': [v['nom'] for v in variantes],
    'traitements': traitements,
    'versions': versions,
    'fichiers': fichiers_crees[:10]  # Limiter à 10 pour le JSON
}

resume_file = os.path.join(output_dir, 'resume.json')
with open(resume_file, 'w', encoding='utf-8') as f:
    json.dump(resume, f, indent=2, ensure_ascii=False)

print(f"\n" + "=" * 80)
print(f"✅ {len(fichiers_crees)} fichiers créés dans {output_dir}")
print("=" * 80)

print(f"\n📋 Résumé:")
print(f"   Largeurs totales: {len(largeurs_totales)}")
print(f"   Profondeurs totales: {len(profondeurs_totales)}")
print(f"   Variantes: {len(variantes)}")
print(f"   Traitements: {len(traitements)}")
print(f"   Versions: {len(versions)}")
print(f"   Total: {len(largeurs_totales)} × {len(profondeurs_totales)} × {len(variantes)} × {len(traitements)} × {len(versions)} = {len(fichiers_crees)} fichiers")

print(f"\n💡 Prochaines étapes:")
print(f"   Utilisez calculateur_prix_camflex.py pour :")
print(f"   1. Calculer automatiquement les formules Excel")
print(f"   2. Extraire les prix et composants")
print(f"   3. Générer le fichier final resultats_tous.json")

