#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Lit les résultats de tous les fichiers Excel dans le dossier résultats
"""

import openpyxl
import json
import os
import glob
from datetime import datetime

print("=" * 80)
print("LECTURE DES RÉSULTATS")
print("=" * 80)

# Trouver tous les fichiers Excel dans le dossier résultats
resultats_dir = 'résultats'
fichiers = []
if os.path.exists(resultats_dir):
    for root, dirs, files in os.walk(resultats_dir):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~'):
                fichiers.append(os.path.join(root, file))

fichiers = sorted(fichiers)

if not fichiers:
    print("❌ Aucun fichier Excel trouvé dans le dossier résultats")
    exit(1)

print(f"\n📦 {len(fichiers)} fichiers trouvés\n")

resultats = []

for fichier in fichiers:
    print(f"📄 Lecture de {fichier}...")
    
    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
        ws_configure = wb['Configure']
        ws_prc = wb['PRC import']
        
        # Lire la configuration
        largeur = ws_configure.cell(1, 2).value
        profondeur = ws_configure.cell(2, 1).value
        treatment = ws_configure.cell(16, 2).value  # B16
        version = ws_configure.cell(17, 2).value  # B17
        
        # Lire le prix
        price_brut = ws_prc.cell(7, 8).value  # H7
        price_remise = ws_prc.cell(8, 8).value  # H8
        price_net = ws_prc.cell(9, 8).value  # H9
        
        # Déterminer le type (ouvert/fermé) depuis le nom du fichier
        fichier_basename = os.path.basename(fichier)
        type_abri = 'inconnu'
        if 'ouvert' in fichier_basename:
            type_abri = 'ouvert'
        elif 'ferme' in fichier_basename:
            type_abri = 'ferme'
        
        # Déterminer la variante
        variante = 'inconnu'
        if 'normal' in fichier_basename:
            variante = 'normal'
        elif 'bosque' in fichier_basename:
            variante = 'bosque'
        
        resultat = {
            'fichier': fichier_basename,
            'chemin_complet': fichier,
            'type': type_abri,
            'variante': variante,
            'largeur': largeur,
            'profondeur': profondeur,
            'treatment': treatment,
            'version': version,
            'prix_brut': price_brut,
            'remise': price_remise,
            'prix_net': price_net
        }
        
        resultats.append(resultat)
        
        if price_net:
            print(f"   ✅ Largeur: {largeur}m, Profondeur: {profondeur}m → Prix: {price_net} €")
        else:
            print(f"   ⚠️  Largeur: {largeur}m, Profondeur: {profondeur}m → Prix non calculé")
            
    except Exception as e:
        print(f"   ❌ Erreur: {e}")

print(f"\n" + "=" * 80)
print("RÉSUMÉ")
print("=" * 80)

# Afficher les résultats
# Sauvegarder les résultats dans un fichier JSON
resultats_file = os.path.join(resultats_dir, 'tous_les_resultats.json')
with open(resultats_file, 'w', encoding='utf-8') as f:
    json.dump({
        'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'total': len(resultats),
        'resultats': resultats
    }, f, indent=2, ensure_ascii=False)

print(f"\n💾 Résultats sauvegardés dans: {resultats_file}")

# Afficher un résumé
print(f"\n{'Fichier':<50} {'Type':>8} {'Variante':>10} {'Largeur':>8} {'Treatment':>12} {'Version':>10} {'Prix net':>12}")
print("-" * 120)

for r in resultats:
    prix = f"{r['prix_net']:.2f} €" if r['prix_net'] else "Non calculé"
    treatment = r.get('treatment', 'N/A')
    version = r.get('version', 'N/A')
    type_abri = r.get('type', 'N/A')
    variante = r.get('variante', 'N/A')
    print(f"{r['fichier']:<50} {type_abri:>8} {variante:>10} {r['largeur']:>8} {treatment:>12} {version:>10} {prix:>12}")

# Sauvegarder en JSON
result_file = 'resultats_tous.json'
with open(result_file, 'w', encoding='utf-8') as f:
    json.dump({
        'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'resultats': resultats
    }, f, indent=2, ensure_ascii=False)

print(f"\n💾 Résultats sauvegardés dans: {result_file}")

# Compter les prix calculés
prix_calcules = [r for r in resultats if r['prix_net'] is not None]
print(f"\n📊 {len(prix_calcules)}/{len(resultats)} fichiers avec prix calculé")

if len(prix_calcules) < len(resultats):
    print(f"\n⚠️  {len(resultats) - len(prix_calcules)} fichiers n'ont pas de prix calculé")
    print(f"   Ouvrez-les dans Excel et appuyez sur F9, puis relancez ce script")

