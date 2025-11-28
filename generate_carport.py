#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère les fichiers Excel pour les abris vélos CARPORTS
- Pas de murs (B21-B24 = "No")
- B19 = "No wall"
- Pas de portes (lignes 28-31 vides)
- Largeurs : 2.5, 4, 5, 6, 7, 8, 9, 10, 11, 12
- Versions : Standard (N) / PLUS (P)
- Traitements : Galvanized (G) / Powder coated (PT)
- Format : CAR-{largeur}M-{version}-{treatment}
"""

import openpyxl
import shutil
import os
import json
from datetime import datetime

# Dossier et fichier source
base_dir = 'fichier de base'
source_file = os.path.join(base_dir, 'fichier_de_prix_de_base.xlsx')
resultats_dir = 'résultats'

# Valeurs valides depuis Excel (feuille Calc)
traitements = ['Galvanized', 'Powder coated']
versions = ['Standard', 'PLUS']

# Largeurs pour Carport
largeurs_totales = [2.5, 4, 5, 6, 7, 8, 9, 10, 11, 12]

# Profondeurs pour Carport (comme les ouverts : 2m et 2.5m)
profondeurs_totales = [2, 2.5]

# Fonction pour décomposer la largeur en valeurs valides (2m, 2.5m, 4m, 5m, 6m)
def decomposer_largeur(largeur_totale):
    """
    Décompose une largeur totale en valeurs valides (2.03, 2.53, 4.06, 5.06, 6.09)
    Retourne une liste de valeurs à mettre dans B1, C1, D1, etc.
    """
    valeurs_valides = [2.03, 2.53, 4.06, 5.06, 6.09]
    
    # Cas spéciaux selon les règles
    if largeur_totale == 2.5:
        return [2.53]
    elif largeur_totale == 4:
        return [4.06]
    elif largeur_totale == 5:
        return [5.06]
    elif largeur_totale == 6:
        return [6.09]
    elif largeur_totale == 7:
        return [2.53, 2.03, 2.53]  # 2.5m + 2m + 2.5m
    elif largeur_totale == 8:
        return [4.06, 4.06]  # 4m + 4m
    elif largeur_totale == 9:
        return [2.53, 4.06, 2.53]  # 2.5m + 4m + 2.5m
    elif largeur_totale == 10:
        return [5.06, 5.06]  # 5m + 5m
    elif largeur_totale == 11:
        return [2.53, 6.09, 2.53]  # 2.5m + 6m + 2.5m
    elif largeur_totale == 12:
        return [6.09, 6.09]  # 6m + 6m
    else:
        # Algorithme glouton pour les autres cas
        resultat = []
        reste = largeur_totale
        valeurs_triees = sorted(valeurs_valides, reverse=True)
        
        while reste > 0.1:
            trouve = False
            for val in valeurs_triees:
                if reste >= val:
                    resultat.append(val)
                    reste -= val
                    trouve = True
                    break
            if not trouve:
                resultat.append(valeurs_valides[0])
                reste = 0
        
        return resultat

print("=" * 80)
print("GÉNÉRATION DES ABRIS VÉLOS CARPORTS")
print("=" * 80)

# Vérifier que le fichier source existe
if not os.path.exists(source_file):
    print(f"❌ Erreur: {source_file} n'existe pas !")
    exit(1)

# Créer le dossier résultats
os.makedirs(resultats_dir, exist_ok=True)

# Créer un sous-dossier pour les carports
output_dir = os.path.join(resultats_dir, 'carport')
os.makedirs(output_dir, exist_ok=True)

# Supprimer les anciens fichiers
for old_file in os.listdir(output_dir):
    if old_file.endswith('.xlsx'):
        os.remove(os.path.join(output_dir, old_file))

fichiers_crees = []
compteur = 1

# Générer toutes les combinaisons
for largeur_totale in largeurs_totales:
    # Décomposer en valeurs valides
    largeurs_decomposees = decomposer_largeur(largeur_totale)
    
    for profondeur_totale in profondeurs_totales:
        # Pour les carports, la profondeur est directement 2.03m ou 2.53m
        if profondeur_totale == 2:
            profondeur_decomposee = [2.03]
        elif profondeur_totale == 2.5:
            profondeur_decomposee = [2.53]
        else:
            profondeur_decomposee = [2.03]  # Par défaut
        
        for treatment in traitements:
            for version in versions:
                # Nom du fichier selon la nomenclature
                # Format: CAR-{largeur}M-{version}-{profondeur}-{treatment}
                
                # Type: CAR pour Carport
                type_code = 'CAR'
                
                # Largeur: format 2.5M, 4M, 10M, etc.
                if largeur_totale == int(largeur_totale):
                    largeur_code = f'{int(largeur_totale)}M'
                else:
                    largeur_code = f'{largeur_totale}M'
                
                # Version: N pour Standard, P pour PLUS
                version_code = 'N' if version == 'Standard' else 'P'
                
                # Profondeur: format 200 pour 2m, 250 pour 2.5m
                profondeur_code = str(int(profondeur_totale * 100))
                
                # Treatment: G pour Galvanized, PT pour Powder coated
                treatment_code = 'G' if treatment == 'Galvanized' else 'PT'
                
                # Nom du fichier
                nom_fichier = f'{type_code}-{largeur_code}-{version_code}-{profondeur_code}-{treatment_code}.xlsx'
                work_file = os.path.join(output_dir, nom_fichier)
                
                print(f"\n📦 Création {compteur}: {os.path.basename(work_file)}")
                print(f"   Largeur totale: {largeur_totale}m → {largeurs_decomposees}")
                print(f"   Profondeur totale: {profondeur_totale}m → {profondeur_decomposee}")
                
                # Dupliquer le fichier source
                shutil.copy2(source_file, work_file)
                
                # Ouvrir et modifier
                wb = openpyxl.load_workbook(work_file, data_only=False)
                ws = wb['Configure']
                
                # Nettoyer les lignes 28-31 (supprimer les espaces, zéros et valeurs, mettre à None pour les ouverts)
                for row in range(28, 32):
                    for col in range(1, 4):  # Colonnes A, B, C
                        cell_value = ws.cell(row, col).value
                        # Pour les ouverts, toutes les cellules doivent être vides
                        if cell_value is not None:
                            ws.cell(row, col).value = None
                
                # Mettre "*" dans toutes les cellules de dimensions
                for row in range(2, 14):
                    ws.cell(row, 1).value = "*"
                for col in range(2, 8):
                    ws.cell(1, col).value = "*"
                
                # Écrire la profondeur décomposée (A2)
                ws.cell(2, 1).value = profondeur_decomposee[0]  # A2 = profondeur (2.03 ou 2.53)
                
                # Écrire les largeurs décomposées (B1, C1, D1, etc.)
                for i, larg in enumerate(largeurs_decomposees[:6]):  # Max 6 largeurs
                    ws.cell(1, 2 + i).value = larg
                
                # Écrire les options de base
                ws.cell(16, 2).value = treatment  # B16 = treatment
                ws.cell(17, 2).value = version  # B17 = version
                
                # Configuration CARPORT
                ws.cell(19, 2).value = 'No wall'  # B19 = "No wall"
                ws.cell(21, 2).value = 'No'  # B21 = top wall (No)
                ws.cell(22, 2).value = 'No'  # B22 = right wall (No)
                ws.cell(23, 2).value = 'No'  # B23 = bottom wall (No)
                ws.cell(24, 2).value = 'No'  # B24 = left wall (No)
                ws.cell(25, 2).value = 'No'  # B25 = remove cladding
                
                # Pas de portes pour les carports - Les lignes 28-31 sont déjà nettoyées (None)
                # A28, B28, C28, A29, B29, C29, A30, B30, C30, A31, B31, C31 = None (vides)
                
                # Nettoyer A33 et B33 (cellules fusionnées) - doivent être vides pour les carports
                # Démerger si nécessaire, puis vider
                merged_ranges = list(ws.merged_cells.ranges)
                for merged_range in merged_ranges:
                    if merged_range.min_row == 33 and merged_range.min_col <= 2 and merged_range.max_col >= 2:
                        ws.unmerge_cells(str(merged_range))
                        break
                ws.cell(33, 1).value = None  # A33
                ws.cell(33, 2).value = None  # B33
                
                # NE PAS TOUCHER B26 et B27 - ils ne doivent pas être modifiés
                
                # Sauvegarder
                wb.save(work_file)
                
                fichiers_crees.append({
                    'fichier': os.path.basename(work_file),
                    'largeur_totale': largeur_totale,
                    'largeurs_decomposees': largeurs_decomposees,
                    'profondeur_totale': profondeur_totale,
                    'profondeur_decomposee': profondeur_decomposee,
                    'treatment': treatment,
                    'version': version,
                    'type': 'carport'
                })
                
                compteur += 1

# Sauvegarder le résumé
resume = {
    'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    'type': 'carport',
    'total_fichiers': len(fichiers_crees),
    'largeurs_totales': largeurs_totales,
    'profondeurs_totales': profondeurs_totales,
    'traitements': traitements,
    'versions': versions,
    'fichiers': fichiers_crees[:10]  # Limiter à 10 pour le JSON
}

resume_file = os.path.join(output_dir, 'resume.json')
with open(resume_file, 'w', encoding='utf-8') as f:
    json.dump(resume, f, indent=2, ensure_ascii=False)

print(f"\n" + "=" * 80)
print(f"✅ {len(fichiers_crees)} fichiers créés dans {output_dir}")
print("=" * 80)

print(f"\n📋 Résumé:")
print(f"   Largeurs: {len(largeurs_totales)}")
print(f"   Profondeurs: {len(profondeurs_totales)}")
print(f"   Traitements: {len(traitements)}")
print(f"   Versions: {len(versions)}")
print(f"   Total: {len(largeurs_totales)} × {len(profondeurs_totales)} × {len(traitements)} × {len(versions)} = {len(fichiers_crees)} fichiers")

print(f"\n💡 Prochaines étapes:")
print(f"   Utilisez calculateur_prix_camflex.py pour :")
print(f"   1. Calculer automatiquement les formules Excel")
print(f"   2. Extraire les prix et composants")
print(f"   3. Générer le fichier final resultats_tous.json")

