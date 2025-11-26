#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère les fichiers Excel pour les abris MÉTALLIQUES OUVERTS
- B19 = "2D mesh" (wall material)
- B20 = "RAV716" (mesh finish)
- Murs : haut, droite, gauche (pas en bas) - OUVERT
- Pas de portes
- Remove cladding : No (B25 = No)
- Variantes : Standard/PLUS × Galvanized/Powder coated
- Toutes les combinaisons de largeurs et profondeurs
"""

import openpyxl
import shutil
import os
import json
from datetime import datetime

# Dossier et fichier source
base_dir = 'fichier de base'
source_file = os.path.join(base_dir, 'nepastoucher.xlsx')
resultats_dir = 'résultats'

# Valeurs valides depuis Excel (feuille Calc)
traitements = ['Galvanized', 'Powder coated']
versions = ['Standard', 'PLUS']

# Variantes - Pour Métallique, remove_cladding = No
variantes = [
    {'nom': 'metallique', 'wall_material': '2D mesh', 'mesh_finish': 'RAV716', 'remove_cladding': 'No'}
]

# Fonction pour décomposer la profondeur en valeurs valides (2m, 2.5m)
def decomposer_profondeur(profondeur_totale):
    """
    Décompose une profondeur totale en valeurs valides (2.03m, 2.53m)
    Retourne une liste de valeurs à mettre dans A2, A3, A4, etc.
    """
    if profondeur_totale == 4:
        return [2.03, 2.03]
    elif profondeur_totale == 4.5:
        return [2.03, 2.53]
    elif profondeur_totale == 5:
        return [2.53, 2.53]
    elif profondeur_totale == 6:
        return [2.03, 2.03, 2.03]
    elif profondeur_totale == 6.5:
        return [2.03, 2.03, 2.53]
    elif profondeur_totale == 7:
        return [2.03, 2.53, 2.53]
    else:
        valeurs_valides = [2.03, 2.53]
        resultat = []
        reste = profondeur_totale
        
        while reste > 0.1:
            if reste >= 2.53:
                resultat.append(2.53)
                reste -= 2.53
            elif reste >= 2.03:
                resultat.append(2.03)
                reste -= 2.03
            else:
                resultat.append(2.03)
                reste = 0
        
        return resultat

# Fonction pour décomposer la largeur en valeurs valides
def decomposer_largeur(largeur_totale):
    """
    Décompose une largeur totale en valeurs valides
    """
    if largeur_totale == 2:
        return [2.03]
    elif largeur_totale == 2.5:
        return [2.53]
    elif largeur_totale == 4:
        return [4.06]
    elif largeur_totale == 5:
        return [5.06]
    elif largeur_totale == 6:
        return [6.09]
    elif largeur_totale == 7:
        return [2.53, 2.03, 2.53]
    elif largeur_totale == 8:
        return [4.06, 4.06]
    elif largeur_totale == 9:
        return [2.53, 4.06, 2.53]
    elif largeur_totale == 10:
        return [5.06, 5.06]
    elif largeur_totale == 11:
        return [2.53, 6.09, 2.53]
    elif largeur_totale == 12:
        return [6.09, 6.09]
    elif largeur_totale == 13:
        return [4.06, 5.06, 4.06]
    elif largeur_totale == 14:
        return [5.06, 4.06, 5.06]
    else:
        valeurs_valides = [2.03, 2.53, 4.06, 5.06, 6.09]
        resultat = []
        reste = largeur_totale
        
        while reste > 0.1:
            trouve = False
            for val in sorted(valeurs_valides, reverse=True):
                if reste >= val:
                    resultat.append(val)
                    reste -= val
                    trouve = True
                    break
            if not trouve:
                resultat.append(2.03)
                reste = 0
        
        return resultat

# Largeurs et profondeurs pour les ouverts
largeurs_totales = [2, 2.5, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]
profondeurs_totales = [4, 4.5, 5, 6, 6.5, 7]

print("=" * 80)
print("GÉNÉRATION DES ABRIS MÉTALLIQUES OUVERTS")
print("=" * 80)

# Vérifier que le fichier source existe
if not os.path.exists(source_file):
    print(f"❌ Erreur: {source_file} n'existe pas !")
    exit(1)

# Créer le dossier résultats
os.makedirs(resultats_dir, exist_ok=True)

# Créer un sous-dossier pour les abris métalliques ouverts
output_dir = os.path.join(resultats_dir, 'metallique_ouvert')
os.makedirs(output_dir, exist_ok=True)

# Supprimer les anciens fichiers
for old_file in os.listdir(output_dir):
    if old_file.endswith('.xlsx'):
        os.remove(os.path.join(output_dir, old_file))

fichiers_crees = []
compteur = 1

# Générer toutes les combinaisons
for largeur_totale in largeurs_totales:
    largeurs_decomposees = decomposer_largeur(largeur_totale)
    
    for profondeur_totale in profondeurs_totales:
        profondeurs_decomposees = decomposer_profondeur(profondeur_totale)
        
        for variante in variantes:
            for treatment in traitements:
                for version in versions:
                    # Nom du fichier selon la nomenclature
                    # Format: MET-{largeur}M-{version}-{profondeur}-{treatment}
                    
                    type_code = 'MET'
                    
                    # Largeur: format 2M, 2.5M, etc.
                    if largeur_totale == int(largeur_totale):
                        largeur_code = f'{int(largeur_totale)}M'
                    else:
                        largeur_code = f'{largeur_totale}M'
                    
                    # Version: N pour Standard, P pour PLUS
                    version_code = 'N' if version == 'Standard' else 'P'
                    
                    # Profondeur: format 400, 450, 500, etc.
                    profondeur_code = str(int(profondeur_totale * 100))
                    
                    # Treatment: G pour Galvanized, PT pour Powder coated
                    treatment_code = 'G' if treatment == 'Galvanized' else 'PT'
                    
                    # Nom du fichier
                    nom_fichier = f'{type_code}-{largeur_code}-{version_code}-{profondeur_code}-{treatment_code}.xlsx'
                    work_file = os.path.join(output_dir, nom_fichier)
                    
                    print(f"\n📦 Création {compteur}: {os.path.basename(work_file)}")
                    print(f"   Largeur totale: {largeur_totale}m → {largeurs_decomposees}")
                    print(f"   Profondeur totale: {profondeur_totale}m → {profondeurs_decomposees}")
                    
                    # Dupliquer le fichier source
                    shutil.copy2(source_file, work_file)
                    
                    # Ouvrir et modifier
                    wb = openpyxl.load_workbook(work_file, data_only=False)
                    ws = wb['Configure']
                    
                    # Mettre "*" dans toutes les cellules de dimensions
                    for row in range(2, 14):
                        ws.cell(row, 1).value = "*"
                    for col in range(2, 8):
                        ws.cell(1, col).value = "*"
                    
                    # Écrire les profondeurs décomposées (A2, A3, A4, etc.)
                    for i, prof in enumerate(profondeurs_decomposees[:12]):
                        ws.cell(2 + i, 1).value = prof
                    
                    # Écrire les largeurs décomposées (B1, C1, D1, etc.)
                    for i, larg in enumerate(largeurs_decomposees[:6]):
                        ws.cell(1, 2 + i).value = larg
                    
                    # Écrire les options de base
                    ws.cell(16, 2).value = treatment  # B16 = treatment
                    ws.cell(17, 2).value = version  # B17 = version
                    
                    # Configuration MÉTALLIQUE OUVERT
                    ws.cell(19, 2).value = variante['wall_material']  # B19 = "2D mesh"
                    ws.cell(20, 2).value = variante['mesh_finish']  # B20 = "RAV716"
                    ws.cell(21, 2).value = 'Yes'  # B21 = top wall
                    ws.cell(22, 2).value = 'Yes'  # B22 = right wall
                    ws.cell(23, 2).value = 'No'  # B23 = bottom wall (OUVERT)
                    ws.cell(24, 2).value = 'Yes'  # B24 = left wall
                    ws.cell(25, 2).value = 'No'  # B25 = remove cladding
                    
                    # Pas de portes pour les ouverts
                    ws.cell(28, 1).value = None  # A28 = entrance type (vide)
                    ws.cell(28, 2).value = 0
                    ws.cell(28, 3).value = 0
                    
                    # NE PAS TOUCHER B26 et B27 - ils ne doivent pas être modifiés
                    
                    # Sauvegarder
                    wb.save(work_file)
                    
                    fichiers_crees.append({
                        'fichier': os.path.basename(work_file),
                        'largeur_totale': largeur_totale,
                        'profondeur_totale': profondeur_totale,
                        'variante': variante['nom'],
                        'treatment': treatment,
                        'version': version,
                        'type': 'metallique_ouvert'
                    })
                    
                    compteur += 1

# Sauvegarder le résumé
resume = {
    'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    'type': 'metallique_ouvert',
    'total_fichiers': len(fichiers_crees),
    'largeurs_totales': largeurs_totales,
    'profondeurs_totales': profondeurs_totales,
    'variantes': [v['nom'] for v in variantes],
    'traitements': traitements,
    'versions': versions,
    'fichiers': fichiers_crees[:10]
}

resume_file = os.path.join(output_dir, 'resume.json')
with open(resume_file, 'w', encoding='utf-8') as f:
    json.dump(resume, f, indent=2, ensure_ascii=False)

print(f"\n" + "=" * 80)
print(f"✅ {len(fichiers_crees)} fichiers créés dans {output_dir}")
print("=" * 80)

print(f"\n📋 Résumé:")
print(f"   Largeurs: {len(largeurs_totales)}")
print(f"   Profondeurs: {len(profondeurs_totales)}")
print(f"   Variantes: {len(variantes)}")
print(f"   Traitements: {len(traitements)}")
print(f"   Versions: {len(versions)}")
print(f"   Total: {len(largeurs_totales)} × {len(profondeurs_totales)} × {len(variantes)} × {len(traitements)} × {len(versions)} = {len(fichiers_crees)} fichiers")

