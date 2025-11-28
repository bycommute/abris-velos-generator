#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fusionne tous les fichiers Excel générés en un seul fichier Excel
avec des feuilles organisées par type d'abri
"""

import openpyxl
import os
import json
from datetime import datetime
from collections import defaultdict

print("=" * 80)
print("FUSION DE TOUS LES FICHIERS EXCEL")
print("=" * 80)

# Dossier résultats
resultats_dir = 'résultats'
output_file = os.path.join(resultats_dir, 'TOUS_LES_RESULTATS.xlsx')

# Trouver tous les fichiers Excel
fichiers = []
if os.path.exists(resultats_dir):
    for root, dirs, files in os.walk(resultats_dir):
        for file in files:
            if file.endswith('.xlsx') and not file.startswith('~') and file != 'TOUS_LES_RESULTATS.xlsx':
                fichiers.append(os.path.join(root, file))

fichiers = sorted(fichiers)

if not fichiers:
    print("❌ Aucun fichier Excel trouvé dans le dossier résultats")
    exit(1)

print(f"\n📦 {len(fichiers)} fichiers trouvés\n")

# Organiser les fichiers par type
fichiers_par_type = defaultdict(list)

for fichier in fichiers:
    chemin_relatif = os.path.relpath(fichier, resultats_dir)
    dossier = os.path.dirname(chemin_relatif)
    
    # Déterminer le type depuis le chemin
    if 'bosquet_ferme' in dossier:
        type_abri = 'Bosquet Fermé'
        sous_type = 'ferme'
    elif 'bosquet_ferme_compact' in dossier:
        type_abri = 'Bosquet Fermé Compact'
        sous_type = 'ferme_compact'
    elif 'bosquet_ouvert' in dossier:
        type_abri = 'Bosquet Ouvert'
        sous_type = 'ouvert'
    elif 'bosquet_ouvert_compact' in dossier:
        type_abri = 'Bosquet Ouvert Compact'
        sous_type = 'ouvert_compact'
    elif 'domino_ferme' in dossier:
        type_abri = 'Domino Fermé'
        sous_type = 'ferme'
    elif 'domino_ferme_compact' in dossier:
        type_abri = 'Domino Fermé Compact'
        sous_type = 'ferme_compact'
    elif 'domino_ouvert' in dossier:
        type_abri = 'Domino Ouvert'
        sous_type = 'ouvert'
    elif 'domino_ouvert_compact' in dossier:
        type_abri = 'Domino Ouvert Compact'
        sous_type = 'ouvert_compact'
    elif 'metallique_ferme' in dossier:
        type_abri = 'Métallique Fermé'
        sous_type = 'ferme'
    elif 'metallique_ferme_compact' in dossier:
        type_abri = 'Métallique Fermé Compact'
        sous_type = 'ferme_compact'
    elif 'metallique_ouvert' in dossier:
        type_abri = 'Métallique Ouvert'
        sous_type = 'ouvert'
    elif 'metallique_ouvert_compact' in dossier:
        type_abri = 'Métallique Ouvert Compact'
        sous_type = 'ouvert_compact'
    elif 'neve_ferme' in dossier:
        type_abri = 'Neve Fermé'
        sous_type = 'ferme'
    elif 'neve_ferme_compact' in dossier:
        type_abri = 'Neve Fermé Compact'
        sous_type = 'ferme_compact'
    elif 'neve_ouvert' in dossier:
        type_abri = 'Neve Ouvert'
        sous_type = 'ouvert'
    else:
        type_abri = 'Autre'
        sous_type = 'autre'
    
    fichiers_par_type[type_abri].append(fichier)

print(f"📊 Types d'abris trouvés: {len(fichiers_par_type)}\n")
for type_abri, liste_fichiers in sorted(fichiers_par_type.items()):
    print(f"   {type_abri}: {len(liste_fichiers)} fichiers")

# Créer un nouveau workbook
wb_final = openpyxl.Workbook()
wb_final.remove(wb_final.active)  # Supprimer la feuille par défaut

# Créer une feuille de résumé
ws_resume = wb_final.create_sheet("Résumé", 0)

# En-têtes de la feuille résumé
ws_resume['A1'] = 'Type'
ws_resume['B1'] = 'Fichier'
ws_resume['C1'] = 'Largeur (m)'
ws_resume['D1'] = 'Profondeur (m)'
ws_resume['E1'] = 'Treatment'
ws_resume['F1'] = 'Version'
ws_resume['G1'] = 'Prix Brut (€)'
ws_resume['H1'] = 'Remise (€)'
ws_resume['I1'] = 'Prix Net (€)'
ws_resume['J1'] = 'Chemin'

# Style des en-têtes
from openpyxl.styles import Font, PatternFill, Alignment

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws_resume[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

ligne_resume = 2
total_fichiers_traites = 0
erreurs = []

# Traiter chaque type d'abri
for type_abri in sorted(fichiers_par_type.keys()):
    print(f"\n📋 Traitement: {type_abri} ({len(fichiers_par_type[type_abri])} fichiers)")
    
    # Créer une feuille pour ce type (nom limité à 31 caractères pour Excel)
    nom_feuille = type_abri[:31]
    ws_type = wb_final.create_sheet(nom_feuille)
    
    # En-têtes pour la feuille de type
    ws_type['A1'] = 'Fichier'
    ws_type['B1'] = 'Largeur (m)'
    ws_type['C1'] = 'Profondeur (m)'
    ws_type['D1'] = 'Treatment'
    ws_type['E1'] = 'Version'
    ws_type['F1'] = 'Prix Brut (€)'
    ws_type['G1'] = 'Remise (€)'
    ws_type['H1'] = 'Prix Net (€)'
    ws_type['I1'] = 'Chemin'
    
    # Style des en-têtes
    for cell in ws_type[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ligne_type = 2
    
    # Traiter chaque fichier de ce type
    for fichier in sorted(fichiers_par_type[type_abri]):
        try:
            wb = openpyxl.load_workbook(fichier, data_only=True)
            
            # Lire les données
            ws_configure = wb['Configure']
            ws_prc = wb['PRC import']
            
            largeur = ws_configure.cell(1, 2).value
            profondeur = ws_configure.cell(2, 1).value
            treatment = ws_configure.cell(16, 2).value
            version = ws_configure.cell(17, 2).value
            
            price_brut = ws_prc.cell(7, 8).value
            price_remise = ws_prc.cell(8, 8).value
            price_net = ws_prc.cell(9, 8).value
            
            fichier_basename = os.path.basename(fichier)
            chemin_relatif = os.path.relpath(fichier, resultats_dir)
            
            # Ajouter à la feuille de type
            ws_type.cell(ligne_type, 1).value = fichier_basename
            ws_type.cell(ligne_type, 2).value = largeur
            ws_type.cell(ligne_type, 3).value = profondeur
            ws_type.cell(ligne_type, 4).value = treatment
            ws_type.cell(ligne_type, 5).value = version
            ws_type.cell(ligne_type, 6).value = price_brut
            ws_type.cell(ligne_type, 7).value = price_remise
            ws_type.cell(ligne_type, 8).value = price_net
            ws_type.cell(ligne_type, 9).value = chemin_relatif
            
            # Ajouter à la feuille résumé
            ws_resume.cell(ligne_resume, 1).value = type_abri
            ws_resume.cell(ligne_resume, 2).value = fichier_basename
            ws_resume.cell(ligne_resume, 3).value = largeur
            ws_resume.cell(ligne_resume, 4).value = profondeur
            ws_resume.cell(ligne_resume, 5).value = treatment
            ws_resume.cell(ligne_resume, 6).value = version
            ws_resume.cell(ligne_resume, 7).value = price_brut
            ws_resume.cell(ligne_resume, 8).value = price_remise
            ws_resume.cell(ligne_resume, 9).value = price_net
            ws_resume.cell(ligne_resume, 10).value = chemin_relatif
            
            ligne_type += 1
            ligne_resume += 1
            total_fichiers_traites += 1
            
            if total_fichiers_traites % 50 == 0:
                print(f"   ✅ {total_fichiers_traites} fichiers traités...")
                
        except Exception as e:
            erreur_msg = f"Erreur avec {fichier}: {e}"
            erreurs.append(erreur_msg)
            print(f"   ❌ {erreur_msg}")
    
    # Ajuster la largeur des colonnes pour la feuille de type
    ws_type.column_dimensions['A'].width = 40
    ws_type.column_dimensions['B'].width = 12
    ws_type.column_dimensions['C'].width = 12
    ws_type.column_dimensions['D'].width = 15
    ws_type.column_dimensions['E'].width = 12
    ws_type.column_dimensions['F'].width = 15
    ws_type.column_dimensions['G'].width = 15
    ws_type.column_dimensions['H'].width = 15
    ws_type.column_dimensions['I'].width = 50

# Ajuster la largeur des colonnes pour la feuille résumé
ws_resume.column_dimensions['A'].width = 25
ws_resume.column_dimensions['B'].width = 40
ws_resume.column_dimensions['C'].width = 12
ws_resume.column_dimensions['D'].width = 12
ws_resume.column_dimensions['E'].width = 15
ws_resume.column_dimensions['F'].width = 12
ws_resume.column_dimensions['G'].width = 15
ws_resume.column_dimensions['H'].width = 15
ws_resume.column_dimensions['I'].width = 15
ws_resume.column_dimensions['J'].width = 50

# Ajouter un filtre automatique sur la feuille résumé
ws_resume.auto_filter.ref = ws_resume.dimensions

# Sauvegarder le fichier
print(f"\n💾 Sauvegarde du fichier fusionné...")
wb_final.save(output_file)

print(f"\n" + "=" * 80)
print("✅ FUSION TERMINÉE")
print("=" * 80)
print(f"\n📁 Fichier créé: {output_file}")
print(f"📊 Total de fichiers traités: {total_fichiers_traites}")
print(f"📋 Nombre de feuilles: {len(wb_final.sheetnames)}")

if erreurs:
    print(f"\n⚠️  {len(erreurs)} erreurs rencontrées:")
    for erreur in erreurs[:10]:  # Afficher les 10 premières
        print(f"   - {erreur}")
    if len(erreurs) > 10:
        print(f"   ... et {len(erreurs) - 10} autres erreurs")

print(f"\n💡 Instructions:")
print(f"   1. Ouvrez le fichier {output_file}")
print(f"   2. La feuille 'Résumé' contient tous les résultats")
print(f"   3. Les autres feuilles sont organisées par type d'abri")
print(f"   4. Utilisez les filtres pour rechercher facilement")

