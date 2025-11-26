#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère les fichiers Excel pour les abris Domino OUVERTS COMPACT
- Profondeur : toujours 2.5m
- Largeurs : 2m, 2.5m, 4m, 5m, 6m
- Murs : haut, droite, gauche (pas en bas) - OUVERT
- Pas de portes
- Remove cladding : Yes (B25 = Yes)
- Variantes : Standard/PLUS × Galvanized/Powder coated
"""

import openpyxl
import shutil
import os
import json
from datetime import datetime

# Dossier et fichier source
base_dir = 'fichier de base'
source_file = os.path.join(base_dir, 'nepastoucher.xlsx')
resultats_dir = 'résultats'

# Valeurs valides depuis Excel (feuille Calc)
traitements = ['Galvanized', 'Powder coated']
versions = ['Standard', 'PLUS']

# Variantes - Pour Domino, remove_cladding = Yes
variantes = [
    {'nom': 'normal', 'wall_material': 'Thermowood', 'remove_cladding': 'Yes'},
    {'nom': 'bosque', 'wall_material': 'Thermowood', 'remove_cladding': 'Yes'}
]

# Largeurs pour compact
largeurs_totales = [2, 2.5, 4, 5, 6]
profondeur_fixe = 2.5  # Toujours 2.5m

print("=" * 80)
print("GÉNÉRATION DES ABRIS DOMINO OUVERTS COMPACT")
print("=" * 80)

# Vérifier que le fichier source existe
if not os.path.exists(source_file):
    print(f"❌ Erreur: {source_file} n'existe pas !")
    exit(1)

# Créer le dossier résultats
os.makedirs(resultats_dir, exist_ok=True)

# Créer un sous-dossier pour les abris Domino ouverts compact
output_dir = os.path.join(resultats_dir, 'domino_ouvert_compact')
os.makedirs(output_dir, exist_ok=True)

# Supprimer les anciens fichiers
for old_file in os.listdir(output_dir):
    if old_file.endswith('.xlsx'):
        os.remove(os.path.join(output_dir, old_file))

fichiers_crees = []
compteur = 1

# Générer toutes les combinaisons
for largeur_totale in largeurs_totales:
    for variante in variantes:
        for treatment in traitements:
            for version in versions:
                # Nom du fichier selon la nomenclature
                # Format: DOM-COMPACT-{largeur}M-{version}-{treatment}
                
                # Type: DOM-COMPACT pour Domino ouvert compact
                type_code = 'DOM-COMPACT'
                
                # Largeur: format 2M, 2.5M, etc.
                if largeur_totale == int(largeur_totale):
                    largeur_code = f'{int(largeur_totale)}M'
                else:
                    largeur_code = f'{largeur_totale}M'
                
                # Version: N pour Standard, P pour PLUS
                version_code = 'N' if version == 'Standard' else 'P'
                
                # Profondeur: toujours 2.5m = 250
                profondeur_code = '250'
                
                # Treatment: G pour Galvanized, PT pour Powder coated
                treatment_code = 'G' if treatment == 'Galvanized' else 'PT'
                
                # Nom du fichier
                nom_fichier = f'{type_code}-{largeur_code}-{version_code}-{profondeur_code}-{treatment_code}.xlsx'
                work_file = os.path.join(output_dir, nom_fichier)
                
                print(f"\n📦 Création {compteur}: {os.path.basename(work_file)}")
                print(f"   Largeur: {largeur_totale}m | Profondeur: {profondeur_fixe}m")
                
                # Dupliquer le fichier source
                shutil.copy2(source_file, work_file)
                
                # Ouvrir et modifier
                wb = openpyxl.load_workbook(work_file, data_only=False)
                ws = wb['Configure']
                
                # Mettre "*" dans toutes les cellules de dimensions
                for row in range(2, 14):
                    ws.cell(row, 1).value = "*"
                for col in range(2, 8):
                    ws.cell(1, col).value = "*"
                
                # Écrire la profondeur (toujours 2.5m)
                ws.cell(2, 1).value = 2.53  # A2 = 2.5m
                
                # Écrire la largeur (directe, pas de décomposition pour les compacts)
                if largeur_totale == 2:
                    ws.cell(1, 2).value = 2.03  # B1 = 2m
                elif largeur_totale == 2.5:
                    ws.cell(1, 2).value = 2.53  # B1 = 2.5m
                elif largeur_totale == 4:
                    ws.cell(1, 2).value = 4.06  # B1 = 4m (direct)
                elif largeur_totale == 5:
                    ws.cell(1, 2).value = 5.06  # B1 = 5m (direct)
                elif largeur_totale == 6:
                    ws.cell(1, 2).value = 6.09  # B1 = 6m (direct)
                
                # Écrire les options de base
                ws.cell(16, 2).value = treatment  # B16 = treatment
                ws.cell(17, 2).value = version  # B17 = version
                
                # Configuration DOMINO OUVERT COMPACT
                ws.cell(19, 2).value = variante['wall_material']  # B19 = wall material
                ws.cell(21, 2).value = 'Yes'  # B21 = top wall
                ws.cell(22, 2).value = 'Yes'  # B22 = right wall
                ws.cell(23, 2).value = 'No'  # B23 = bottom wall (OUVERT)
                ws.cell(24, 2).value = 'Yes'  # B24 = left wall
                ws.cell(25, 2).value = 'Yes'  # B25 = remove cladding (Yes pour Domino)
                
                # Pas de portes pour les ouverts
                # A28, B28, C28, A33 restent vides (comme dans le fichier de base)
                
                # NE PAS TOUCHER B26 et B27 - ils ne doivent pas être modifiés
                
                # Sauvegarder
                wb.save(work_file)
                
                fichiers_crees.append({
                    'fichier': os.path.basename(work_file),
                    'largeur_totale': largeur_totale,
                    'profondeur_totale': profondeur_fixe,
                    'variante': variante['nom'],
                    'treatment': treatment,
                    'version': version,
                    'type': 'domino_ouvert_compact'
                })
                
                compteur += 1

# Sauvegarder le résumé
resume = {
    'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    'type': 'domino_ouvert_compact',
    'total_fichiers': len(fichiers_crees),
    'largeurs_totales': largeurs_totales,
    'profondeur_fixe': profondeur_fixe,
    'variantes': [v['nom'] for v in variantes],
    'traitements': traitements,
    'versions': versions,
    'fichiers': fichiers_crees[:10]  # Limiter à 10 pour le JSON
}

resume_file = os.path.join(output_dir, 'resume.json')
with open(resume_file, 'w', encoding='utf-8') as f:
    json.dump(resume, f, indent=2, ensure_ascii=False)

print(f"\n" + "=" * 80)
print(f"✅ {len(fichiers_crees)} fichiers créés dans {output_dir}")
print("=" * 80)

print(f"\n📋 Résumé:")
print(f"   Largeurs: {len(largeurs_totales)}")
print(f"   Profondeur fixe: {profondeur_fixe}m")
print(f"   Variantes: {len(variantes)}")
print(f"   Traitements: {len(traitements)}")
print(f"   Versions: {len(versions)}")
print(f"   Total: {len(largeurs_totales)} × {len(variantes)} × {len(traitements)} × {len(versions)} = {len(fichiers_crees)} fichiers")

print(f"\n💡 Instructions:")
print(f"   1. Ouvrez chaque fichier dans Excel")
print(f"   2. Appuyez sur F9 pour recalculer")
print(f"   3. Fermez Excel")
print(f"   4. Utilisez read_results.py pour lire les prix")

