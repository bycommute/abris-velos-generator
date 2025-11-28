#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script pour extraire les prix (H7, H9) et les composants (A2:E110) de tous les fichiers Excel
- Traite uniquement les fichiers sans prix
- Système de retry limité (2 tentatives par run, réinitialisé à chaque lancement)
- Gestion robuste de la mémoire (max 2 workers)
- Sauvegarde fréquente pour éviter la perte de données
"""

import openpyxl
import json
import os
import subprocess
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import time

# Configuration
resultats_dir = 'résultats'
composant_dir = 'composant'
resultats_json_file = 'resultats_tous.json'
max_workers = 2  # Réduit à 2 pour la stabilité (était 5)
max_attempts_per_run = 2  # Maximum 2 tentatives par fichier par run
delay_between_files = 1.5  # Délai entre chaque fichier pour laisser Excel se stabiliser

# Lock pour thread-safe writing
json_lock = Lock()

def load_existing_results():
    """Charge les résultats existants depuis resultats_tous.json"""
    if os.path.exists(resultats_json_file):
        try:
            with open(resultats_json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Créer un dictionnaire indexé par chemin pour vérification rapide
                results_dict = {}
                for r in data.get('resultats', []):
                    chemin = r.get('chemin_complet', '')
                    if chemin:
                        results_dict[chemin] = r
                return data, results_dict
        except Exception as e:
            print(f"⚠️  Erreur lors du chargement de {resultats_json_file}: {e}")
            return {'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'resultats': []}, {}
    return {'date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 'resultats': []}, {}

def save_results(data):
    """Sauvegarde les résultats dans resultats_tous.json"""
    with json_lock:
        try:
            # Trier les résultats par chemin pour un ordre cohérent
            data['resultats'] = sorted(data['resultats'], key=lambda x: x.get('chemin_complet', ''))
            data['date_derniere_maj'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            data['total'] = len(data['resultats'])
            
            with open(resultats_json_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"❌ Erreur lors de la sauvegarde: {e}")

def extract_components(ws_prc):
    """Extrait les composants de A2 à E110 de la feuille PRC import"""
    components = []
    for row in range(2, 111):  # Lignes 2 à 110
        row_data = []
        for col in range(1, 6):  # Colonnes A à E (1 à 5)
            cell_value = ws_prc.cell(row, col).value
            # Convertir les valeurs pour JSON
            if cell_value is None:
                row_data.append(None)
            elif isinstance(cell_value, (int, float)):
                row_data.append(cell_value)
            else:
                row_data.append(str(cell_value))
        components.append(row_data)
    return components

def open_and_calculate_excel(file_path):
    """
    Ouvre un fichier Excel dans Excel, force le recalcul, sauvegarde et ferme.
    Utilise AppleScript pour contrôler Excel sur macOS.
    Délais augmentés pour laisser le temps aux formules complexes de se calculer.
    """
    file_path_abs = os.path.abspath(file_path)
    
    # Script AppleScript amélioré avec délais plus longs
    applescript = f'''
    tell application "Microsoft Excel"
        -- Vérifier que le fichier n'est pas déjà ouvert
        set fileRef to POSIX file "{file_path_abs}"
        try
            set wb to workbook (name of fileRef)
            close wb saving no
        end try
        
        -- Ouvrir le fichier
        open fileRef
        
        -- Attendre que le fichier soit chargé
        delay 0.8
        
        -- Forcer le recalcul complet de toutes les formules
        calculate workbook
        
        -- Attendre que les calculs soient terminés (délai augmenté)
        delay 2.5
        
        -- Sauvegarder le fichier
        save active workbook
        
        -- Attendre que la sauvegarde soit terminée
        delay 0.5
        
        -- Fermer le fichier
        close active workbook saving yes
        
        -- Attendre que la fermeture soit terminée
        delay 0.3
    end tell
    '''
    
    try:
        result = subprocess.run(
            ['osascript', '-e', applescript],
            capture_output=True,
            text=True,
            timeout=45  # Timeout augmenté
        )
        
        if result.returncode == 0:
            return True, None
        else:
            error_msg = result.stderr.strip() if result.stderr else "Erreur inconnue"
            return False, error_msg
            
    except subprocess.TimeoutExpired:
        return False, "Timeout: le fichier a pris trop de temps"
    except Exception as e:
        return False, str(e)

def is_valid_price(value):
    """
    Vérifie si une valeur est un prix valide (nombre > 0)
    """
    if value is None:
        return False
    if not isinstance(value, (int, float)):
        return False
    if value <= 0:
        return False
    return True

def process_excel_file(file_path, existing_results_dict, attempt_number):
    """
    Traite un fichier Excel et extrait les prix et composants.
    attempt_number: numéro de la tentative (1 ou 2)
    """
    fichier_basename = os.path.basename(file_path)
    
    # Déterminer le type d'abri depuis le chemin
    type_abri = 'autre'
    if 'carport' in file_path:
        type_abri = 'carport'
    elif 'bosquet_ferme' in file_path:
        type_abri = 'bosquet_ferme'
    elif 'bosquet_ouvert' in file_path:
        type_abri = 'bosquet_ouvert'
    elif 'domino_ferme' in file_path:
        type_abri = 'domino_ferme'
    elif 'domino_ouvert' in file_path:
        type_abri = 'domino_ouvert'
    elif 'metallique_ferme' in file_path:
        type_abri = 'metallique_ferme'
    elif 'metallique_ouvert' in file_path:
        type_abri = 'metallique_ouvert'
    elif 'neve_ouvert' in file_path:
        type_abri = 'neve_ouvert'
    
    try:
        # ÉTAPE 1 : Ouvrir le fichier dans Excel pour calculer les formules
        success, error = open_and_calculate_excel(file_path)
        if not success:
            return None, f"Erreur ouverture Excel: {error}", False
        
        # Délai pour laisser Excel se stabiliser
        time.sleep(0.5)
        
        # ÉTAPE 2 : Lire les données calculées
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        if 'PRC import' not in wb.sheetnames:
            wb.close()
            return None, f"Feuille 'PRC import' introuvable", False
        
        ws_prc = wb['PRC import']
        
        # Lire les prix avec vérification
        prix_avant_raw = ws_prc.cell(7, 8).value  # H7
        prix_apres_raw = ws_prc.cell(9, 8).value  # H9
        
        # Vérifier que les prix sont valides (nombres > 0)
        prix_avant = prix_avant_raw if is_valid_price(prix_avant_raw) else None
        prix_apres = prix_apres_raw if is_valid_price(prix_apres_raw) else None
        
        # Extraire les composants (A2:E110)
        components = extract_components(ws_prc)
        
        wb.close()
        
        # Créer le résultat
        result = {
            'fichier': fichier_basename,
            'chemin_complet': file_path,
            'type_abri': type_abri,
            'prix_avant_reduction': prix_avant,
            'prix_apres_reduction': prix_apres,
            'date_extraction': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'tentative': attempt_number
        }
        
        # Sauvegarder les composants dans un fichier JSON séparé
        composant_type_dir = os.path.join(composant_dir, type_abri)
        os.makedirs(composant_type_dir, exist_ok=True)
        
        composant_file = os.path.join(composant_type_dir, fichier_basename.replace('.xlsx', '.json'))
        
        # Vérifier si le fichier composant existe déjà
        composants_existants = None
        if os.path.exists(composant_file):
            try:
                with open(composant_file, 'r', encoding='utf-8') as f:
                    data_existant = json.load(f)
                    composants_existants = data_existant.get('composants', [])
            except:
                pass
        
        # Vérifier si les nouveaux composants sont valides (non vides)
        composants_valides = components and len(components) > 0 and any(
            any(cell is not None and cell != '' for cell in row) 
            for row in components
        )
        
        # Écraser seulement si les nouveaux composants sont valides
        # Sinon, garder l'ancien s'il existe
        if composants_valides or composants_existants is None:
            composant_data = {
                'fichier_source': fichier_basename,
                'chemin_source': file_path,
                'date_extraction': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'composants': components
            }
            
            with json_lock:
                with open(composant_file, 'w', encoding='utf-8') as f:
                    json.dump(composant_data, f, indent=2, ensure_ascii=False)
        # Si les nouveaux composants sont vides mais qu'il y a un ancien, on garde l'ancien
        elif composants_existants:
            # Mettre à jour seulement la date d'extraction mais garder les anciens composants
            composant_data = {
                'fichier_source': fichier_basename,
                'chemin_source': file_path,
                'date_extraction': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'composants': composants_existants,
                'note': 'Composants conservés (nouvelle extraction vide)'
            }
            
            with json_lock:
                with open(composant_file, 'w', encoding='utf-8') as f:
                    json.dump(composant_data, f, indent=2, ensure_ascii=False)
        
        return result, None, False
        
    except Exception as e:
        return None, f"Erreur: {e}", False

def find_excel_files(directory):
    """Trouve tous les fichiers Excel dans le dossier résultats"""
    fichiers = []
    if os.path.exists(directory):
        for root, dirs, files in os.walk(directory):
            for file in files:
                if file.endswith('.xlsx') and not file.startswith('~') and file != 'TOUS_LES_RESULTATS.xlsx':
                    fichiers.append(os.path.join(root, file))
    return sorted(fichiers)

def get_type_abri_from_path(file_path):
    """Détermine le type d'abri depuis le chemin"""
    if 'carport' in file_path:
        return 'carport'
    elif 'bosquet_ferme' in file_path:
        return 'bosquet_ferme'
    elif 'bosquet_ouvert' in file_path:
        return 'bosquet_ouvert'
    elif 'domino_ferme' in file_path:
        return 'domino_ferme'
    elif 'domino_ouvert' in file_path:
        return 'domino_ouvert'
    elif 'metallique_ferme' in file_path:
        return 'metallique_ferme'
    elif 'metallique_ouvert' in file_path:
        return 'metallique_ouvert'
    elif 'neve_ouvert' in file_path:
        return 'neve_ouvert'
    return 'autre'

def main():
    print("=" * 80)
    print("EXTRACTION DES PRIX ET COMPOSANTS (VERSION AMÉLIORÉE)")
    print("=" * 80)
    
    # Charger les résultats existants
    print("\n📖 Chargement des résultats existants...")
    results_data, existing_results_dict = load_existing_results()
    print(f"   {len(existing_results_dict)} fichiers déjà dans les résultats")
    
    # Trouver tous les fichiers Excel
    print("\n🔍 Recherche des fichiers Excel...")
    fichiers = find_excel_files(resultats_dir)
    print(f"   {len(fichiers)} fichiers trouvés")
    
    if not fichiers:
        print("❌ Aucun fichier Excel trouvé")
        return
    
    # Créer le dossier composant
    os.makedirs(composant_dir, exist_ok=True)
    
    # Activer Excel une seule fois au début
    print("🔧 Activation d'Excel...")
    subprocess.run(['osascript', '-e', 'tell application "Microsoft Excel" to activate'], 
                   capture_output=True)
    time.sleep(1)
    
    # Filtrer les fichiers : ne traiter que ceux SANS PRIX
    fichiers_a_traiter = []
    fichiers_avec_prix = 0
    fichiers_sans_resultat = 0
    
    for fichier in fichiers:
        # Vérifier si le fichier a déjà un prix dans les résultats
        if fichier in existing_results_dict:
            resultat = existing_results_dict[fichier]
            prix_avant = resultat.get('prix_avant_reduction')
            prix_apres = resultat.get('prix_apres_reduction')
            
            # Si le fichier a des prix valides, on ne le retraite pas
            if is_valid_price(prix_avant) and is_valid_price(prix_apres):
                fichiers_avec_prix += 1
                continue
            # Sinon, on le retraite (pas de prix ou prix invalides)
            else:
                fichiers_a_traiter.append(fichier)
        else:
            # Fichier pas encore dans les résultats, on le traite
            fichiers_sans_resultat += 1
            fichiers_a_traiter.append(fichier)
    
    print(f"\n📊 Fichiers à traiter (sans prix): {len(fichiers_a_traiter)}")
    print(f"   Fichiers avec prix complets: {fichiers_avec_prix}")
    print(f"   Fichiers sans résultat: {fichiers_sans_resultat}")
    
    if not fichiers_a_traiter:
        print("\n✅ Tous les fichiers ont des prix complets !")
        return
    
    # Système de retry : dictionnaire pour suivre les tentatives pendant ce run
    attempts_dict = {}  # {file_path: attempt_count}
    
    # Préparer la liste des fichiers avec leurs tentatives
    files_with_attempts = []
    for fichier in fichiers_a_traiter:
        attempts_dict[fichier] = 0
        files_with_attempts.append((fichier, 1))  # Première tentative
    
    # Traiter les fichiers en parallèle
    print(f"\n🚀 Traitement en parallèle avec {max_workers} workers...")
    print(f"   Maximum {max_attempts_per_run} tentatives par fichier par run")
    print()
    
    succes = 0
    echecs = 0
    fichiers_sans_prix_final = []
    start_time = time.time()
    completed = 0
    
    # Créer une copie locale des résultats pour mise à jour
    results_list = results_data.get('resultats', [])
    results_dict_local = existing_results_dict.copy()
    
    def process_with_retry(file_path, attempt_num):
        """Traite un fichier avec gestion du retry"""
        nonlocal results_list, results_dict_local, attempts_dict
        
        # Incrémenter le compteur de tentatives
        attempts_dict[file_path] = attempt_num
        
        result, error, deja_fait = process_excel_file(file_path, results_dict_local, attempt_num)
        
        if result:
            prix_avant = result.get('prix_avant_reduction')
            prix_apres = result.get('prix_apres_reduction')
            
            # Vérifier si on a obtenu des prix valides
            if is_valid_price(prix_avant) and is_valid_price(prix_apres):
                # Succès : on a des prix valides
                chemin = result['chemin_complet']
                if chemin in results_dict_local:
                    # Mettre à jour l'existant
                    idx = next((i for i, r in enumerate(results_list) if r.get('chemin_complet') == chemin), None)
                    if idx is not None:
                        results_list[idx].update(result)
                else:
                    # Ajouter nouveau
                    results_list.append(result)
                    results_dict_local[chemin] = result
                
                return file_path, result, True, None, None
            else:
                # Pas de prix valides, on peut retenter si on n'a pas atteint le max
                if attempt_num < max_attempts_per_run:
                    return file_path, None, False, None, "Prix non valides, retry possible"
                else:
                    return file_path, None, False, None, "Prix non valides après 2 tentatives"
        else:
            # Erreur lors du traitement
            if attempt_num < max_attempts_per_run:
                return file_path, None, False, error, "Erreur, retry possible"
            else:
                return file_path, None, False, error, "Erreur après 2 tentatives"
    
    # Traiter les fichiers avec retry
    remaining_files = files_with_attempts.copy()
    
    while remaining_files:
        # Préparer les fichiers pour ce round
        current_round = []
        for file_path, attempt_num in remaining_files:
            if attempt_num <= max_attempts_per_run:
                current_round.append((file_path, attempt_num))
        
        if not current_round:
            break
        
        # Traiter ce round en parallèle
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {
                executor.submit(process_with_retry, file_path, attempt_num): (file_path, attempt_num)
                for file_path, attempt_num in current_round
            }
            
            next_round = []
            
            for future in as_completed(future_to_file):
                file_path, attempt_num = future_to_file[future]
                try:
                    file_path_result, result, success, error, status = future.result()
                    completed += 1
                    basename = os.path.basename(file_path_result)
                    
                    if success:
                        prix_avant = result.get('prix_avant_reduction')
                        prix_apres = result.get('prix_apres_reduction')
                        print(f"[{completed}] ✅ {basename} | Tentative {attempt_num}/{max_attempts_per_run} | Avant: {prix_avant:.2f} € | Après: {prix_apres:.2f} €")
                        succes += 1
                        
                        # Sauvegarder immédiatement après chaque succès
                        results_data['resultats'] = results_list
                        save_results(results_data)
                        
                        # Délai entre fichiers
                        time.sleep(delay_between_files)
                    else:
                        # Vérifier si on peut retenter
                        if attempt_num < max_attempts_per_run and "retry possible" in status:
                            # Retenter
                            next_round.append((file_path_result, attempt_num + 1))
                            print(f"[{completed}] ⚠️  {basename} | Tentative {attempt_num}/{max_attempts_per_run} | {status}")
                        else:
                            # Plus de tentatives possibles
                            fichiers_sans_prix_final.append(file_path_result)
                            error_msg = error if error else status
                            print(f"[{completed}] ❌ {basename} | Tentative {attempt_num}/{max_attempts_per_run} | {error_msg}")
                            echecs += 1
                            
                            # Sauvegarder quand même (même sans prix)
                            chemin = file_path_result
                            if chemin in results_dict_local:
                                idx = next((i for i, r in enumerate(results_list) if r.get('chemin_complet') == chemin), None)
                                if idx is not None:
                                    results_list[idx]['date_extraction'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                type_abri = get_type_abri_from_path(chemin)
                                result_no_price = {
                                    'fichier': os.path.basename(chemin),
                                    'chemin_complet': chemin,
                                    'type_abri': type_abri,
                                    'prix_avant_reduction': None,
                                    'prix_apres_reduction': None,
                                    'date_extraction': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    'tentative': attempt_num,
                                    'erreur': error_msg
                                }
                                results_list.append(result_no_price)
                                results_dict_local[chemin] = result_no_price
                            
                            # Sauvegarder périodiquement (tous les 5 fichiers)
                            if completed % 5 == 0:
                                results_data['resultats'] = results_list
                                save_results(results_data)
                            
                            # Délai entre fichiers
                            time.sleep(delay_between_files)
                    
                except Exception as e:
                    completed += 1
                    basename = os.path.basename(file_path)
                    print(f"[{completed}] ❌ {basename} | Exception: {e}")
                    echecs += 1
                    time.sleep(delay_between_files)
        
        # Mettre à jour la liste des fichiers restants
        remaining_files = next_round
    
    # Sauvegarder les résultats finaux
    results_data['resultats'] = results_list
    save_results(results_data)
    
    elapsed_time = time.time() - start_time
    
    # Résumé
    print(f"\n" + "=" * 80)
    print("RÉSUMÉ")
    print("=" * 80)
    print(f"✅ Succès (avec prix): {succes}")
    if echecs > 0:
        print(f"❌ Échecs (sans prix après {max_attempts_per_run} tentatives): {echecs}")
    print(f"⏱️  Temps total: {elapsed_time:.1f} secondes")
    if len(fichiers_a_traiter) > 0:
        print(f"⚡ Temps moyen par fichier: {elapsed_time/len(fichiers_a_traiter):.1f} secondes")
    
    print(f"\n💾 Résultats sauvegardés dans: {resultats_json_file}")
    print(f"💾 Composants sauvegardés dans: {composant_dir}/")
    
    # Statistiques sur les prix
    prix_complets = [r for r in results_list if is_valid_price(r.get('prix_avant_reduction')) and is_valid_price(r.get('prix_apres_reduction'))]
    print(f"\n📊 {len(prix_complets)}/{len(results_list)} fichiers avec prix complets")
    
    # Liste des fichiers sans prix après plusieurs tentatives
    if fichiers_sans_prix_final:
        print(f"\n⚠️  FICHIERS SANS PRIX APRÈS {max_attempts_per_run} TENTATIVES ({len(fichiers_sans_prix_final)} fichiers):")
        for fichier in fichiers_sans_prix_final[:20]:  # Afficher les 20 premiers
            print(f"   - {os.path.basename(fichier)}")
        if len(fichiers_sans_prix_final) > 20:
            print(f"   ... et {len(fichiers_sans_prix_final) - 20} autres")
        print(f"\n💡 Ces fichiers seront retraités au prochain run (2 nouvelles tentatives)")

if __name__ == '__main__':
    main()
