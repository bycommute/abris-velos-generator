#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur d'URLs SharePoint Drive
==================================

Ce script parcourt le dossier "résultats" et génère un tableau CSV/Excel
avec les URLs SharePoint Drive pour chaque fichier, en suivant la logique :

- Base dossiers : https://camflexsystems.sharepoint.com/:f:/r/sites/agentportal/ByCommute/Domino%20Tool/Tous_les_variants_bycommute/
- Base fichiers : https://camflexsystems.sharepoint.com/:x:/r/sites/agentportal/ByCommute/Domino%20Tool/Tous_les_variants_bycommute/
- URL dossier : {base_dossiers}{nom_dossier}?web=1
- URL fichier : {base_fichiers}{nom_dossier}/{nom_fichier}?web=1

Utilisation :
    python generate_drive_urls.py
"""

import os
import csv
from pathlib import Path
from urllib.parse import quote

# Configuration
RESULTATS_DIR = 'résultats'
OUTPUT_CSV = 'urls_drive.csv'
OUTPUT_EXCEL = 'urls_drive.xlsx'

# Bases d'URL SharePoint
BASE_DOSSIERS = 'https://camflexsystems.sharepoint.com/:f:/r/sites/agentportal/ByCommute/Domino%20Tool/Tous_les_variants_bycommute/'
BASE_FICHIERS = 'https://camflexsystems.sharepoint.com/:x:/r/sites/agentportal/ByCommute/Domino%20Tool/Tous_les_variants_bycommute/'


def encoder_url(texte):
    """
    Encode un texte pour une URL (équivalent de ENCODEURL en Excel)
    Gère les espaces, accents, caractères spéciaux, etc.
    Les underscores (_) restent tels quels si possible.
    """
    # Utiliser quote avec safe='' pour encoder tous les caractères spéciaux
    # mais on peut garder certains caractères comme safe si nécessaire
    return quote(texte, safe='')


def generer_url_dossier(nom_dossier):
    """Génère l'URL SharePoint pour un dossier"""
    nom_encode = encoder_url(nom_dossier)
    return f"{BASE_DOSSIERS}{nom_encode}?web=1"


def generer_url_fichier(nom_dossier, nom_fichier):
    """Génère l'URL SharePoint pour un fichier"""
    dossier_encode = encoder_url(nom_dossier)
    fichier_encode = encoder_url(nom_fichier)
    return f"{BASE_FICHIERS}{dossier_encode}/{fichier_encode}?web=1"


def parcourir_resultats():
    """
    Parcourt le dossier résultats et collecte tous les fichiers
    Retourne une liste de tuples (nom_dossier, nom_fichier)
    """
    fichiers = []
    
    if not os.path.exists(RESULTATS_DIR):
        print(f"❌ Le dossier '{RESULTATS_DIR}' n'existe pas")
        return fichiers
    
    print(f"📁 Parcours du dossier '{RESULTATS_DIR}'...\n")
    
    # Parcourir tous les sous-dossiers
    for item in os.listdir(RESULTATS_DIR):
        chemin_sous_dossier = os.path.join(RESULTATS_DIR, item)
        
        # Vérifier que c'est bien un dossier
        if os.path.isdir(chemin_sous_dossier):
            nom_dossier = item
            print(f"   📂 {nom_dossier}")
            
            # Parcourir les fichiers dans ce sous-dossier
            for fichier in os.listdir(chemin_sous_dossier):
                chemin_fichier = os.path.join(chemin_sous_dossier, fichier)
                
                # Ne prendre que les fichiers (pas les sous-dossiers)
                if os.path.isfile(chemin_fichier):
                    # Ignorer les fichiers temporaires Excel
                    if not fichier.startswith('~') and not fichier.startswith('.'):
                        fichiers.append((nom_dossier, fichier))
                        print(f"      📄 {fichier}")
    
    return fichiers


def generer_csv(fichiers):
    """Génère un fichier CSV avec les URLs"""
    print(f"\n💾 Génération du fichier CSV : {OUTPUT_CSV}")
    
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        
        # En-têtes
        writer.writerow(['Nom du dossier', 'Nom du fichier', 'URL du fichier'])
        
        # Données
        for nom_dossier, nom_fichier in fichiers:
            url_fichier = generer_url_fichier(nom_dossier, nom_fichier)
            
            writer.writerow([nom_dossier, nom_fichier, url_fichier])
    
    print(f"   ✅ CSV généré : {len(fichiers)} fichiers traités")


def generer_excel(fichiers):
    """Génère un fichier Excel avec les URLs (nécessite openpyxl)"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        print(f"\n💾 Génération du fichier Excel : {OUTPUT_EXCEL}")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "URLs Drive"
        
        # Style pour les en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # En-têtes
        headers = ['Nom du dossier', 'Nom du fichier', 'URL du fichier']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Bases d'URL en E1 et E2 (pour référence)
        ws['E1'] = 'Base fichiers'
        ws['E2'] = BASE_FICHIERS
        
        # Style pour les bases d'URL
        ws['E1'].font = Font(bold=True)
        
        # Données
        for row, (nom_dossier, nom_fichier) in enumerate(fichiers, start=2):
            url_fichier = generer_url_fichier(nom_dossier, nom_fichier)
            
            ws.cell(row=row, column=1, value=nom_dossier)
            ws.cell(row=row, column=2, value=nom_fichier)
            ws.cell(row=row, column=3, value=url_fichier)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 100
        ws.column_dimensions['E'].width = 15
        
        wb.save(OUTPUT_EXCEL)
        print(f"   ✅ Excel généré : {len(fichiers)} fichiers traités")
        
    except ImportError:
        print(f"   ⚠️  openpyxl n'est pas installé. Installation du CSV uniquement.")
        print(f"   💡 Pour générer Excel, installez : pip install openpyxl")


def main():
    """Fonction principale"""
    print("=" * 80)
    print("GÉNÉRATEUR D'URLS SHAREPOINT DRIVE")
    print("=" * 80)
    print()
    
    # Parcourir le dossier résultats
    fichiers = parcourir_resultats()
    
    if not fichiers:
        print("\n❌ Aucun fichier trouvé dans le dossier résultats")
        return
    
    print(f"\n📊 Total : {len(fichiers)} fichiers trouvés")
    
    # Générer les fichiers de sortie
    generer_csv(fichiers)
    generer_excel(fichiers)
    
    print("\n" + "=" * 80)
    print("✅ GÉNÉRATION TERMINÉE")
    print("=" * 80)
    print(f"\n📄 Fichiers générés :")
    print(f"   • {OUTPUT_CSV}")
    if os.path.exists(OUTPUT_EXCEL):
        print(f"   • {OUTPUT_EXCEL}")
    print(f"\n💡 Les URLs sont prêtes à être utilisées pour héberger les fichiers sur SharePoint Drive.")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠️  Processus interrompu par l'utilisateur")
    except Exception as e:
        print(f"\n\n❌ Erreur : {e}")
        import traceback
        traceback.print_exc()

